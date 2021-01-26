import openpyxl

wb= openpyxl.Workbook()
sheet1 = wb['Sheet']
sheet1.title = '파이선데이터 입력'
sheet1['A1'] = '첫번째 시트'

sheet2 = wb.create_sheet('크롤링 데이터')
sheet2.cell(row=1, column=1).value ='두번째 크롤링'

sheet1.append(['다시', '첫번째 시트'])
wb.save('test2.xlsx')