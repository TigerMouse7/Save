import requests
from bs4 import BeautifulSoup
import openpyxl
from urllib.parse import quote_plus
import datetime
from fake_useragent import UserAgent

now = datetime.datetime.now()
nowDate = now.strftime('%Y%m%d')


excel_file = openpyxl.Workbook()
excel_sheet = excel_file.active
excel_sheet.column_dimensions['A'].width = 100
excel_sheet.column_dimensions['B'].width = 100
num2 = 0
num = 0
pageNum = 1

n = int(input("몇페이지 가져올까요 : "))
lastPage = int(n)

excel_sheet.append(['번호', '제목'])

while pageNum < lastPage:
    url = f'https://news.naver.com/main/list.nhn?mode=LS2D&sid2=258&sid1=101&mid=shm&date={nowDate}&page='
    html = requests.get(url + str(pageNum))
    soup = BeautifulSoup(html.text, features='html.parser')

    products = soup.select('#main_content > div.list_body.newsflash_body > ul.type06 > li > dl > dt')

    for item in products:
        num += 1
        print(item.text.strip())

        excel_sheet.cell(row=num, column=1).value= (item.text)

    pageNum += 1

url2='https://search.shopping.naver.com/best100v2/detail.nhn?catId=50000002&listType=B10002'
html2=requests.get(url2).text
soup2=BeautifulSoup(html2, 'html.parser')

products2 = soup2.select('#productListArea > ul > li > div.thumb_area > a > img')

for item2 in products2:
    print(item2.get("alt"));
    num2 += 1
    excel_sheet.cell(row=num2, column=2, value= item2.get("alt"))

cell_A1 = excel_sheet['A1']
cell_A1.alignment = openpyxl.styles.Alignment(horizontal="center")

cell_B1 = excel_sheet['B1']
cell_B1.alignment = openpyxl.styles.Alignment(horizontal="center")

excel_file.save('school.xlsx')
excel_file.close()